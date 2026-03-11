# Handles loading and validating input data from JSON files.
"""Input/output helpers for loading JSON and building model objects."""

import json
from pathlib import Path

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from models import Class, Course, Room, Teacher


def load_data(file_path):
    """
    Loads data from a JSON file and validates its structure.

    Args:
        file_path: The path to the JSON file containing the input data.
    Returns:
        dict: A dictionary containing the loaded and validated data.
    Raises:
        ValueError: If the JSON file is malformed or missing required fields.
    """
    try:
        path = Path(file_path)
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        if "Classes" not in data:
            raise ValueError("Missing key: Classes")
        if "Courses" not in data:
            raise ValueError("Missing key: Courses")
        if "Teachers" not in data:
            raise ValueError("Missing key: Teachers")
        if "Rooms" not in data:
            raise ValueError("Missing key: Rooms")
        if "Timeslots" not in data:
            raise ValueError("Missing key: Timeslots")

        return data
    except Exception as e:
        raise ValueError(f"Error loading JSON: {e}")


def build_models(data):
    """Converts raw JSON data into internal model objects.

    Args:
        data: Parsed JSON dictionary from `input.json`.

    Returns:
        A dictionary containing all built objects and lookup maps.
    """
    classes = []
    classes_by_name = {}

    for class_data in data["Classes"]:
        class_name = class_data["name"]
        student_count = class_data["student_count"]
        class_obj = Class(class_name, student_count)
        classes.append(class_obj)
        classes_by_name[class_name] = class_obj

    courses = []
    for course_data in data["Courses"]:
        lessons_per_week = 0
        if "lessons_per_week" in course_data:
            lessons_per_week = course_data["lessons_per_week"]
        elif "courses_per_weak" in course_data:
            lessons_per_week = course_data["courses_per_weak"]
        else:
            raise ValueError(f"Course '{course_data.get('name')}' misses lessons key")

        class_names = []
        if "classes" in course_data:
            for class_name in course_data["classes"]:
                class_names.append(class_name)

        course_obj = Course(course_data["name"], lessons_per_week, class_names)
        courses.append(course_obj)

    teachers = []
    for teacher_data in data["Teachers"]:
        abbreviation = None
        if "abbreviation" in teacher_data:
            abbreviation = teacher_data["abbreviation"]

        teacher_obj = Teacher(
            teacher_data["name"],
            teacher_data["courses"],
            abbreviation=abbreviation,
        )
        teachers.append(teacher_obj)

    rooms = []
    for room_data in data["Rooms"]:
        accepted_courses = []
        if "accepted_courses" in room_data:
            for accepted_course in room_data["accepted_courses"]:
                accepted_courses.append(accepted_course)

        room_obj = Room(room_data["name"], room_data["capacity"], accepted_courses)
        rooms.append(room_obj)

    return {
        "classes": classes,
        "classes_by_name": classes_by_name,
        "courses": courses,
        "teachers": teachers,
        "rooms": rooms,
        "timeslots": data["Timeslots"],
    }


def _weekday_rank(day_name):
    order = {
        "Monday": 1,
        "Tuesday": 2,
        "Wednesday": 3,
        "Thursday": 4,
        "Friday": 5,
        "Saturday": 6,
        "Sunday": 7,
    }
    if day_name in order:
        return order[day_name]
    return 999


def _parse_time_start(time_text):
    start_text = time_text
    if "-" in time_text:
        start_text = time_text.split("-")[0]

    try:
        parsed = datetime.strptime(start_text, "%H:%M")
        return parsed.hour * 60 + parsed.minute
    except ValueError:
        return 9999


def _teacher_short(session):
    if session.teacher is None:
        return "-"
    if session.teacher.abbreviation is not None:
        if session.teacher.abbreviation.strip() != "":
            return session.teacher.abbreviation
    return session.teacher.name


def _cell_text_for_lesson(session, room):
    teacher_short = _teacher_short(session)
    return f"{session.course.name} ({teacher_short}) [{room.name}]"


def _build_schedule_lookup(schedule):
    lesson_lookup = {}

    for session, assignment in schedule.items():
        timeslot = assignment[0]
        room = assignment[1]
        key = (session.class_.name, timeslot.day, timeslot.time)
        lesson_lookup[key] = _cell_text_for_lesson(session, room)

    return lesson_lookup


def _collect_day_time_structure(timeslots):
    day_to_times = {}
    for timeslot in timeslots:
        if timeslot.day not in day_to_times:
            day_to_times[timeslot.day] = []

        already_exists = False
        for known_time in day_to_times[timeslot.day]:
            if known_time == timeslot.time:
                already_exists = True
                break

        if not already_exists:
            day_to_times[timeslot.day].append(timeslot.time)

    for day_name in day_to_times:
        day_to_times[day_name].sort(key=_parse_time_start)

    day_names = list(day_to_times.keys())
    day_names.sort(key=_weekday_rank)

    all_times = []
    for day_name in day_names:
        for time_text in day_to_times[day_name]:
            if time_text not in all_times:
                all_times.append(time_text)
    all_times.sort(key=_parse_time_start)

    return day_names, day_to_times, all_times


def _base_styles():
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    return {
        "border": thin_border,
        "title_fill": PatternFill(fill_type="solid", fgColor="D9E1F2"),
        "title_font": Font(size=14, bold=True),
        "header_font": Font(color="FFFFFF", bold=True),
        "header_day_fill": PatternFill(fill_type="solid", fgColor="1F4E78"),
        "header_time_fill": PatternFill(fill_type="solid", fgColor="244062"),
        "free_fill": PatternFill(fill_type="solid", fgColor="FFF2CC"),
        "n_a_fill": PatternFill(fill_type="solid", fgColor="EDEDED"),
        "lesson_fill_a": PatternFill(fill_type="solid", fgColor="F9FBFD"),
        "lesson_fill_b": PatternFill(fill_type="solid", fgColor="FFFFFF"),
    }


def _write_overview_sheet(workbook, lesson_lookup, class_names, day_names, day_to_times, all_times, styles):
    sheet = workbook.active
    sheet.title = "Wochenübersicht"

    total_columns = 1 + (len(day_names) * len(class_names))
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    sheet.cell(row=1, column=1, value="Gesamtansicht: Tag + Klasse je Spalte")
    sheet.cell(row=1, column=1).font = styles["title_font"]
    sheet.cell(row=1, column=1).fill = styles["title_fill"]
    sheet.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

    sheet.cell(row=3, column=1, value="Timeslot")
    sheet.cell(row=3, column=1).font = styles["header_font"]
    sheet.cell(row=3, column=1).fill = styles["header_time_fill"]
    sheet.cell(row=3, column=1).alignment = Alignment(horizontal="center", vertical="center")
    sheet.cell(row=3, column=1).border = styles["border"]

    column = 2
    for day_name in day_names:
        for class_name in class_names:
            header_text = f"{day_name} | {class_name}"
            header_cell = sheet.cell(row=3, column=column, value=header_text)
            header_cell.font = styles["header_font"]
            header_cell.fill = styles["header_day_fill"]
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            header_cell.border = styles["border"]
            column = column + 1

    row = 4
    for index, time_text in enumerate(all_times):
        time_cell = sheet.cell(row=row, column=1, value=time_text)
        time_cell.fill = styles["header_time_fill"]
        time_cell.font = styles["header_font"]
        time_cell.alignment = Alignment(horizontal="center", vertical="center")
        time_cell.border = styles["border"]

        column = 2
        for day_name in day_names:
            for class_name in class_names:
                key = (class_name, day_name, time_text)
                lesson_value = ""

                day_has_time = time_text in day_to_times[day_name]
                if day_has_time:
                    if key in lesson_lookup:
                        lesson_value = lesson_lookup[key]
                    else:
                        lesson_value = "Freistunde"
                else:
                    lesson_value = "—"

                lesson_cell = sheet.cell(row=row, column=column, value=lesson_value)
                lesson_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                lesson_cell.border = styles["border"]

                if lesson_value == "Freistunde":
                    lesson_cell.fill = styles["free_fill"]
                elif lesson_value == "—":
                    lesson_cell.fill = styles["n_a_fill"]
                else:
                    if index % 2 == 0:
                        lesson_cell.fill = styles["lesson_fill_a"]
                    else:
                        lesson_cell.fill = styles["lesson_fill_b"]

                column = column + 1

        row = row + 1

    sheet.freeze_panes = "B4"
    sheet.auto_filter.ref = f"A3:{_excel_column_letter(total_columns)}{row - 1}"
    sheet.column_dimensions["A"].width = 16

    for col in range(2, total_columns + 1):
        col_letter = _excel_column_letter(col)
        sheet.column_dimensions[col_letter].width = 23


def _excel_column_letter(index):
    letters = ""
    value = index
    while value > 0:
        value, rest = divmod(value - 1, 26)
        letters = chr(65 + rest) + letters
    return letters


def _write_class_sheet(workbook, class_name, lesson_lookup, day_names, day_to_times, all_times, styles):
    safe_sheet_name = class_name
    if len(safe_sheet_name) > 31:
        safe_sheet_name = safe_sheet_name[:31]

    sheet = workbook.create_sheet(title=safe_sheet_name)
    total_columns = 1 + len(day_names)

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    sheet.cell(row=1, column=1, value=f"Wochenansicht {class_name}")
    sheet.cell(row=1, column=1).font = styles["title_font"]
    sheet.cell(row=1, column=1).fill = styles["title_fill"]
    sheet.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

    sheet.cell(row=3, column=1, value="Timeslot")
    sheet.cell(row=3, column=1).font = styles["header_font"]
    sheet.cell(row=3, column=1).fill = styles["header_time_fill"]
    sheet.cell(row=3, column=1).alignment = Alignment(horizontal="center", vertical="center")
    sheet.cell(row=3, column=1).border = styles["border"]

    column = 2
    for day_name in day_names:
        header_cell = sheet.cell(row=3, column=column, value=day_name)
        header_cell.font = styles["header_font"]
        header_cell.fill = styles["header_day_fill"]
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        header_cell.border = styles["border"]
        column = column + 1

    row = 4
    for index, time_text in enumerate(all_times):
        time_cell = sheet.cell(row=row, column=1, value=time_text)
        time_cell.fill = styles["header_time_fill"]
        time_cell.font = styles["header_font"]
        time_cell.alignment = Alignment(horizontal="center", vertical="center")
        time_cell.border = styles["border"]

        column = 2
        for day_name in day_names:
            key = (class_name, day_name, time_text)

            if time_text in day_to_times[day_name]:
                if key in lesson_lookup:
                    value = lesson_lookup[key]
                else:
                    value = "Freistunde"
            else:
                value = "—"

            cell = sheet.cell(row=row, column=column, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = styles["border"]

            if value == "Freistunde":
                cell.fill = styles["free_fill"]
            elif value == "—":
                cell.fill = styles["n_a_fill"]
            else:
                if index % 2 == 0:
                    cell.fill = styles["lesson_fill_a"]
                else:
                    cell.fill = styles["lesson_fill_b"]

            column = column + 1

        row = row + 1

    sheet.freeze_panes = "B4"
    sheet.auto_filter.ref = f"A3:{_excel_column_letter(total_columns)}{row - 1}"
    sheet.column_dimensions["A"].width = 16

    for col in range(2, total_columns + 1):
        col_letter = _excel_column_letter(col)
        sheet.column_dimensions[col_letter].width = 26


def export_schedule_to_excel(schedule, output_path, timeslots, classes):
    """Exports the timetable in matrix form (overview + one sheet per class).

    Overview sheet:
    - one column for each (day, class) pair
    - one row per timeslot label

    Class sheet:
    - one column per day
    - one row per timeslot label

    Free periods are explicitly marked as `Freistunde`.
    """
    if schedule is None:
        return

    workbook = Workbook()
    styles = _base_styles()

    class_names = []
    for class_obj in classes:
        class_names.append(class_obj.name)
    class_names.sort()

    lesson_lookup = _build_schedule_lookup(schedule)
    day_names, day_to_times, all_times = _collect_day_time_structure(timeslots)

    _write_overview_sheet(
        workbook,
        lesson_lookup,
        class_names,
        day_names,
        day_to_times,
        all_times,
        styles,
    )

    for class_name in class_names:
        _write_class_sheet(
            workbook,
            class_name,
            lesson_lookup,
            day_names,
            day_to_times,
            all_times,
            styles,
        )

    workbook.save(output_path)


# Example usage
if __name__ == "__main__":
    data = load_data("data/input.json")
    model_data = build_models(data)
    print(f"Classes: {len(model_data['classes'])}")
    print(f"Courses: {len(model_data['courses'])}")
