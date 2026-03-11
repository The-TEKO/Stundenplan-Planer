# Entry point of the application that coordinates input, solving and output.
"""Main workflow of the timetable planner.

This module orchestrates all steps:
- load input data
- create sessions and domains
- run the backtracking solver
- output results to console and Excel
"""

from pathlib import Path
import importlib.util

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from constraints import teacher_for_course
from models import Session
from solver.backtracking import backtracking_search
from timeslot import generate_timeslots
from debug import debug


def _load_project_io_module():
    """Loads the local project `io.py` module explicitly.

    Background:
    Python has a built-in module called `io`.
    This function guarantees that the project's own `io.py` is loaded,
    not the built-in module.
    """
    io_path = Path(__file__).resolve().with_name("io.py")
    spec = importlib.util.spec_from_file_location("project_io", io_path)
    if spec is None:
        raise RuntimeError("Could not create import spec for io.py")
    if spec.loader is None:
        raise RuntimeError("Could not load io.py")

    project_io = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(project_io)
    return project_io


def create_sessions(courses, classes_by_name, teachers) -> list:
    """
    Creates session objects based on the Courses.
    These Sessions then have to be assigned to time slots and rooms while respecting the constraints.
    
    Args:       
        courses: The list of courses to create sessions for.
        classes_by_name: Mapping from class name to class object.
        teachers: Teacher list used to resolve session teacher.

    Returns:
        list: A list of session objects created from the courses.
    """

    sessions = []
    for course in courses:
        teacher = teacher_for_course(course.name, teachers)

        for class_name in course.class_names:
            if class_name not in classes_by_name:
                continue

            class_obj = classes_by_name[class_name]

            lesson_number = 1
            while lesson_number <= course.lessons_per_week:
                session = Session(
                    course=course,
                    class_=class_obj,
                    teacher=teacher,
                    lesson_number=lesson_number,
                )
                sessions.append(session)
                lesson_number = lesson_number + 1

    return sessions

def create_domains(sessions, timeslots, rooms) -> list:
    """
    Creates all possible time slot and room combinations for the sessions. This will be used as the domain for the constraint satisfaction problem.

    Args:
        sessions: The list of session objects for which to create the domains.
        timeslots: The list of available time slots.
        rooms: The list of available rooms.

    Returns:
        list: A dictionary with session as key and list[(timeslot, room)] as value.
    """

    domains = {}

    for session in sessions:
        domains[session] = []

        for timeslot in timeslots:
            for room in rooms:
                if room.capacity < session.class_.student_count:
                    continue

                if len(room.accepted_courses) > 0:
                    accepted = False
                    for accepted_course in room.accepted_courses:
                        if accepted_course == session.course.name:
                            accepted = True
                            break
                    if not accepted:
                        continue

                domains[session].append((timeslot, room))

    return domains


def _sort_assignments_key(item):
    """Sort key for stable, readable console output."""
    session = item[0]
    assignment = item[1]
    timeslot = assignment[0]
    return (session.class_.name, timeslot.day, timeslot.index_in_day, timeslot.time, session.course.name)


def _sort_class_row_key(item, day_order):
    """Sort key for class-specific rows in the Excel export."""
    session = item[0]
    assignment = item[1]
    timeslot = assignment[0]
    return (
        day_order.get(timeslot.day, 999),
        timeslot.index_in_day,
        timeslot.time,
        session.course.name,
    )


def _first_tuple_value(item):
    """Helper for explicit sorting without anonymous functions."""
    return item[0]


def print_schedule(schedule):
    """Prints the generated timetable to the console."""
    if schedule is None:
        print("Keine Lösung gefunden.")
        return

    print("Lösung gefunden. Belegte Sessions:")
    sorted_items = list(schedule.items())
    sorted_items.sort(key=_sort_assignments_key)

    for session, assignment in sorted_items:
        timeslot = assignment[0]
        room = assignment[1]
        teacher_name = "-"
        if session.teacher is not None:
            teacher_name = session.teacher.name

        print(
            f"{session.class_.name} | {timeslot.day} {timeslot.time} | "
            f"{session.course.name} (Lektion {session.lesson_number}) | "
            f"Raum {room.name} | Lehrperson {teacher_name}"
        )


def export_schedule_to_excel(schedule, output_path):
    """Exports the solution as a formatted Excel workbook.

    Structure:
    - one `Overview` sheet with all lessons
    - one sheet per class
    - formatted headers, borders, filters, and frozen header row
    """
    if schedule is None:
        return

    workbook = Workbook()
    overview_sheet = workbook.active
    overview_sheet.title = "Übersicht"

    day_order = {
        "Monday": 1,
        "Tuesday": 2,
        "Wednesday": 3,
        "Thursday": 4,
        "Friday": 5,
        "Saturday": 6,
        "Sunday": 7,
    }

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    title_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    title_font = Font(size=14, bold=True)

    sorted_items = list(schedule.items())
    sorted_items.sort(key=_sort_assignments_key)

    overview_sheet.merge_cells("A1:G1")
    overview_sheet["A1"] = "Stundenplan"
    overview_sheet["A1"].font = title_font
    overview_sheet["A1"].fill = title_fill
    overview_sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Klasse", "Tag", "Zeit", "Fach", "Lektion", "Raum", "Lehrperson"]
    header_row = 3
    column_index = 1
    for header in headers:
        cell = overview_sheet.cell(row=header_row, column=column_index, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        column_index = column_index + 1

    row = 4
    class_names = set()
    for session, assignment in sorted_items:
        timeslot = assignment[0]
        room = assignment[1]
        teacher_name = "-"
        if session.teacher is not None:
            teacher_name = session.teacher.name

        class_names.add(session.class_.name)

        values = [
            session.class_.name,
            timeslot.day,
            timeslot.time,
            session.course.name,
            session.lesson_number,
            room.name,
            teacher_name,
        ]

        col = 1
        for value in values:
            cell = overview_sheet.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            col = col + 1
        row = row + 1

    overview_sheet.freeze_panes = "A4"
    overview_sheet.auto_filter.ref = f"A3:G{row - 1}"
    overview_sheet.column_dimensions["A"].width = 12
    overview_sheet.column_dimensions["B"].width = 14
    overview_sheet.column_dimensions["C"].width = 16
    overview_sheet.column_dimensions["D"].width = 20
    overview_sheet.column_dimensions["E"].width = 10
    overview_sheet.column_dimensions["F"].width = 10
    overview_sheet.column_dimensions["G"].width = 24

    sorted_class_names = list(class_names)
    sorted_class_names.sort()

    for class_name in sorted_class_names:
        safe_sheet_name = class_name
        if len(safe_sheet_name) > 31:
            safe_sheet_name = safe_sheet_name[:31]

        class_sheet = workbook.create_sheet(title=safe_sheet_name)
        class_sheet.merge_cells("A1:F1")
        class_sheet["A1"] = f"Stundenplan {class_name}"
        class_sheet["A1"].font = title_font
        class_sheet["A1"].fill = title_fill
        class_sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

        class_headers = ["Tag", "Zeit", "Fach", "Lektion", "Raum", "Lehrperson"]
        header_col = 1
        for class_header in class_headers:
            header_cell = class_sheet.cell(row=3, column=header_col, value=class_header)
            header_cell.font = header_font
            header_cell.fill = header_fill
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            header_cell.border = thin_border
            header_col = header_col + 1

        class_rows = []
        for session, assignment in sorted_items:
            if session.class_.name == class_name:
                class_rows.append((session, assignment))

        class_rows_with_keys = []
        for class_row in class_rows:
            sort_key = _sort_class_row_key(class_row, day_order)
            class_rows_with_keys.append((sort_key, class_row))

        class_rows_with_keys.sort(key=_first_tuple_value)

        class_rows = []
        for key_and_row in class_rows_with_keys:
            class_rows.append(key_and_row[1])

        current_row = 4
        for session, assignment in class_rows:
            timeslot = assignment[0]
            room = assignment[1]
            teacher_name = "-"
            if session.teacher is not None:
                teacher_name = session.teacher.name

            row_values = [
                timeslot.day,
                timeslot.time,
                session.course.name,
                session.lesson_number,
                room.name,
                teacher_name,
            ]

            current_col = 1
            for row_value in row_values:
                row_cell = class_sheet.cell(row=current_row, column=current_col, value=row_value)
                row_cell.border = thin_border
                row_cell.alignment = Alignment(vertical="center")
                current_col = current_col + 1

            current_row = current_row + 1

        class_sheet.freeze_panes = "A4"
        if current_row > 4:
            class_sheet.auto_filter.ref = f"A3:F{current_row - 1}"

        class_sheet.column_dimensions["A"].width = 14
        class_sheet.column_dimensions["B"].width = 16
        class_sheet.column_dimensions["C"].width = 20
        class_sheet.column_dimensions["D"].width = 10
        class_sheet.column_dimensions["E"].width = 10
        class_sheet.column_dimensions["F"].width = 24

    workbook.save(output_path)


def main():
    """Runs the full planning process and writes the Excel output."""
    project_io = _load_project_io_module()

    project_root = Path(__file__).resolve().parent.parent
    input_path = project_root / "input.json"

    data = project_io.load_data(input_path)
    model_data = project_io.build_models(data)

    timeslots = generate_timeslots(timeslot_definitions=model_data["timeslots"])
    sessions = create_sessions(
        courses=model_data["courses"],
        classes_by_name=model_data["classes_by_name"],
        teachers=model_data["teachers"],
    )
    domains = create_domains(sessions, timeslots, model_data["rooms"])

    debug(f"Timeslots: {len(timeslots)}")
    debug(f"Sessions: {len(sessions)}")

    solution = backtracking_search(sessions, domains, model_data["teachers"])
    print_schedule(solution)

    output_excel = project_root / "output_schedule.xlsx"
    export_schedule_to_excel(solution, output_excel)
    print(f"Excel export erstellt: {output_excel}")

if __name__ == "__main__":
    main()